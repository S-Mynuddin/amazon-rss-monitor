"""Parse Amazon SP-API RSS export and classify each item.

Priority order (most operationally actionable first):
  1. breaking_change      - existing behavior will stop working
  2. deprecation_notice   - retirement announced, future date
  3. schema_change        - fields/attributes/enums changing
  4. security_notice      - auth/credentials/permissions
  5. new_endpoint         - new capability shipped
  6. policy_change        - fees/policy/requirements (non-technical)
  7. regional_availability- market-specific rollouts
  8. informational        - release notes / doc launches / newsletters
  9. other                - catchall

BP-impact rating:
  High - directly touches BP's FBA replenishment pipeline (US/CA marketplaces,
         inventory reports, restock/FBA/orders, product types, listings)
  Med  - adjacent APIs BP uses infrequently or SP-API infrastructure
  Low  - unrelated to BP pipeline (fashion, brand analytics, ads)
  None - marketplaces BP doesn't sell in (EU-only, Japan-only, India-only, Brazil, Mexico)
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ===== UPDATED PATHS FOR YOUR FOLDER STRUCTURE =====
SRC = Path(r"E:\ANDREW WORK PLACE\RSS Feed\Full_email_report.txt")
OUT_DIR = Path(r"E:\ANDREW WORK PLACE\RSS Feed\Output")
# ===================================================

OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_XLSX = OUT_DIR / "SP-API_RSS_categorized.xlsx"


# -------------------------------------------------------------------
# Parser
# -------------------------------------------------------------------
def parse_items(text: str) -> list[dict]:
    items = []
    for chunk in re.split(r"\nITEM #\d+\s*\n", "\n" + text):
        chunk = chunk.strip()
        if not chunk or "Title:" not in chunk:
            continue
        m_title = re.search(r"^Title:\s*(.*?)(?=\n[A-Z][a-z]+ ?[A-Za-z]*:|\Z)", chunk, re.DOTALL | re.MULTILINE)
        m_date  = re.search(r"Published Date:\s*(.*?)(?=\n[A-Z][a-z]+ ?[A-Za-z]*:|\Z)", chunk, re.DOTALL)
        m_link  = re.search(r"Link:\s*(\S+)", chunk)
        m_desc  = re.search(r"Description:\s*(.*?)\Z", chunk, re.DOTALL)
        if not m_title:
            continue
        items.append({
            "title": _clean(m_title.group(1)),
            "date":  _clean(m_date.group(1)) if m_date else "",
            "link":  m_link.group(1).strip() if m_link else "",
            "description": _clean(m_desc.group(1)) if m_desc else "",
        })
    return items


def _clean(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


# -------------------------------------------------------------------
# Classifier (rule-based, priority-ordered)
# -------------------------------------------------------------------

# Compiled regex helpers
BREAKING_RX = re.compile(
    r"\b(will (no longer|be removed|end|be sunset|be turned off|stop (working|being)|be replaced by)|"
    r"no longer offer|no longer support|will cease|is (being )?(retired|removed|sunset)|"
    r"final release|breaking change|action required|must migrate|"
    r"end of (support|life)|will be deprecated|final version|"
    r"will not be available|shall no longer)",
    re.I,
)

DEPRECATION_RX = re.compile(
    r"\b(deprecat(ed|ion|ing)|sunset(ting)?|retiring|will be retired|"
    r"end[- ]of[- ]life|EOL|removal of|removed from the api|will retire)",
    re.I,
)

SCHEMA_RX = re.compile(
    r"\b(attribute usage|enumeration values?|product type definitions?|"
    r"listing attribute|new (field|attribute|parameter|column|response|value|enum|metric|property|dataset|event|list|group-by|element|report type|analytics report)|"
    r"additional (field|attribute|parameter|response|value|metric)|"
    r"schema( change| update)?|response format|payload change|title structure|"
    r"XSD|response body|"
    r"json schema|extended (with|to include)|includes? new|expand(s|ed|ing) (the )?(schema|response|attributes)|"
    r"(now |will )(include|includes|contain|contains|return|returns|support|supports|accept|accepts|show|shows|surface|surfaces|has|have)|"
    r"response now|operation now|(API|report|dataset|feed|endpoint) now|"
    r"format (change|update|standardization)|"
    r"new fee (type|typ)|new value(s)? for|"
    r"rate limit (change|update|increase|adjust)|throttling adjust|"
    r"(attribute|field|property|parameter|object) (added|is now (optional|required)|has been added)|"
    r"(will require|requires|required for|no longer requires)|"
    r"has been added to|"
    r"adds the |adds a new |"
    r"validation of|validate |"
    r"clarification on|"
    r"upcoming (product type|attribute))",
    re.I,
)

SECURITY_RX = re.compile(
    r"\b(OAuth|LWA|restricted data token|\bRDT\b|PII|"
    r"authentication|access token|refresh token|credential|"
    r"grantless|IAM|AWS Signature|encryption|"
    r"permissions? (change|required|update)|authorization|"
    r"security (update|advisory|patch|change)|role scope|application role|"
    r"vulnerabilit(y|ies)|CVE-\d|Log4j|"
    r"escaping issue|injection)",
    re.I,
)

NEW_ENDPOINT_RX = re.compile(
    r"\b(launch(ed|ing)|introduc(ing|es)|announcing|"
    r"new (API|operation|endpoint|version|capability|SDK|sample|documentation|dashboard|tool|workflow)|"
    r"general availability|\bGA release\b|now available|"
    r"prebuilt|new: |Health Dashboard|"
    r"enhancement(s)? to (the )?[A-Z][A-Za-z ]+API|"
    r"Send to Amazon will replace|will replace the|"
    r"released|beta|preview|v\d{4}-\d{2}(-\d{2})?|v\d+_\d+|"
    r"is (now )?live|goes live|rollout|roll[- ]out)",
    re.I,
)

CORRECTION_RX = re.compile(r"\b(correction|clarification|reminder|previously announced|we (announced|previously))", re.I)


POLICY_RX = re.compile(
    r"\b(policy|policies|terms of service|terms and conditions|"
    r"seller (agreement|policy|policies|performance)|"
    r"(FBA|FBM) (requirements?|fees?)|handling time (requirements?|changes?)|"
    r"prep (services?|requirements?|policy)|labeling (services?|requirements?)|"
    r"fee (change|schedule|update|structure)|surcharge|"
    r"compliance|regulatory|requirement (change|update))",
    re.I,
)

REGION_RX = re.compile(
    r"\b(EU marketplace|EU stores|EU sellers|EU FBM|"
    r"United Kingdom|Germany|France|Italy|Spain|Netherlands|Belgium|Poland|Sweden|Turkey|"
    r"Japan|India|Australia|Brazil|Mexico|Singapore|United Arab Emirates|Saudi Arabia|"
    r"South Africa|Egypt|"
    r"marketplace(s)?|region(al)?|country|countries)",
    re.I,
)

INFORMATIONAL_RX = re.compile(
    r"\b(release notes|welcome|documentation site|"
    r"developer newsletter|monthly (update|release|digest)|"
    r"SP-API Release|Selling Partner API Release|Weekly (Update|Digest)|"
    r"quarterly (update|release)|announcement digest|update digest|"
    r"documentation update|blog post|case study)",
    re.I,
)

# BP-impact keyword sets
BP_HIGH_RX = re.compile(
    r"\b(FBA|Fulfillment by Amazon|inventory|restock|replenish|fulfillment|"
    r"Reserved (FC|Customer|Staging)|FC Transfer|"
    r"sales report|orders? report|order( |s )?(api|report)|"
    r"product type|listings? item|catalog item|"
    r"feeds? api|shipment (api|report)|"
    r"inbound (api|shipment)|receiving)",
    re.I,
)

BP_MED_RX = re.compile(
    r"\b(reports? api|pricing (api|report)|notifications|subscription|"
    r"grantless|application management|"
    r"data kiosk|selling partner (api|insights))",
    re.I,
)

BP_LOW_RX = re.compile(
    r"\b(brand analytics|A\+ content|advertising|promotions|solicitations|"
    r"messaging|customer solicitations|invoices?|"
    r"tax(es|ation)|VAT|GST|"
    r"vendor central|1P|first-?party vendor|"
    r"amazon business|B2B|"
    r"eBook|Kindle|Prime Video|automotive fitment|apparel|"
    r"fashion|kitchen|beauty)",
    re.I,
)

# Marketplaces BP is in
BP_MARKETS = re.compile(r"\b(US|USA|United States|CA|Canada|North America|global|worldwide|all marketplaces)\b", re.I)

# Marketplaces BP is NOT in (if EXCLUSIVELY these, drop BP-impact)
NON_BP_MARKETS = re.compile(
    r"\b(EU|Europe|UK|United Kingdom|Germany|France|Italy|Spain|Netherlands|Belgium|Poland|Sweden|Turkey|"
    r"Japan|India|Australia|AU|Brazil|Mexico|Singapore|UAE|United Arab Emirates|Saudi Arabia|"
    r"South Africa|Egypt|Middle East)\b",
    re.I,
)


def classify(item: dict) -> dict:
    """Return {category, secondary_tags, bp_impact, rationale}."""
    haystack = f"{item['title']}\n{item['description']}"
    text_lc = haystack.lower()

    # ---- primary category (priority-ordered) ----
    hits: list[str] = []

    if BREAKING_RX.search(haystack):
        primary = "breaking_change"
        hits.append("breaking")
    elif DEPRECATION_RX.search(haystack):
        primary = "deprecation_notice"
        hits.append("deprecation")
    elif SECURITY_RX.search(haystack):
        primary = "security_notice"
        hits.append("security")
    elif SCHEMA_RX.search(haystack):
        primary = "schema_change"
        hits.append("schema")
    elif NEW_ENDPOINT_RX.search(haystack) and not INFORMATIONAL_RX.search(haystack):
        primary = "new_endpoint"
        hits.append("new")
    elif POLICY_RX.search(haystack):
        primary = "policy_change"
        hits.append("policy")
    elif INFORMATIONAL_RX.search(haystack):
        primary = "informational"
        hits.append("info")
    elif REGION_RX.search(haystack):
        primary = "regional_availability"
        hits.append("regional")
    else:
        primary = "other"
        hits.append("uncategorized")

    # ---- secondary tags: pick up ALL matching signals except the primary ----
    secondary: list[str] = []
    if primary != "breaking_change"  and BREAKING_RX.search(haystack):     secondary.append("breaking")
    if primary != "deprecation_notice" and DEPRECATION_RX.search(haystack): secondary.append("deprecation")
    if primary != "schema_change"    and SCHEMA_RX.search(haystack):        secondary.append("schema")
    if primary != "security_notice"  and SECURITY_RX.search(haystack):      secondary.append("security")
    if primary != "new_endpoint"     and NEW_ENDPOINT_RX.search(haystack):  secondary.append("new")
    if primary != "policy_change"    and POLICY_RX.search(haystack):        secondary.append("policy")
    if primary != "regional_availability" and REGION_RX.search(haystack):   secondary.append("regional")

    # ---- BP-impact rating ----
    has_bp_market  = bool(BP_MARKETS.search(haystack))
    has_non_bp_only = False
    if NON_BP_MARKETS.search(haystack):
        # if the item ONLY names non-BP markets and never names US/CA/global
        has_non_bp_only = not has_bp_market

    # BP has a category-level exclusion for some things: brand analytics, ads, etc.
    is_bp_low_domain = bool(BP_LOW_RX.search(haystack))
    is_bp_high_domain = bool(BP_HIGH_RX.search(haystack))
    is_bp_med_domain = bool(BP_MED_RX.search(haystack))

    # Priority: non-BP market override → domain relevance
    if has_non_bp_only and not is_bp_high_domain:
        bp_impact = "None"
    elif is_bp_high_domain:
        bp_impact = "High"
    elif is_bp_med_domain:
        bp_impact = "Med"
    elif is_bp_low_domain:
        bp_impact = "Low"
    elif primary in ("breaking_change", "deprecation_notice"):
        # even if we can't identify the domain, breaking/deprecation warrants Med by default
        bp_impact = "Med"
    elif primary == "informational":
        bp_impact = "Low"
    else:
        bp_impact = "Med"

    # ---- 1-line rationale ----
    parts = [f"→ {primary}"]
    if secondary:
        parts.append(f"(+ {', '.join(secondary)})")
    if has_non_bp_only:
        parts.append("[non-BP market only]")
    if is_bp_high_domain:
        parts.append("[BP-domain: FBA/inventory/listings]")
    elif is_bp_med_domain:
        parts.append("[BP-domain: reports/notifications]")
    elif is_bp_low_domain:
        parts.append("[BP-domain: ads/brand/vendor — not BP]")
    rationale = " ".join(parts)

    return {
        "category": primary,
        "secondary_tags": ", ".join(secondary),
        "bp_impact": bp_impact,
        "rationale": rationale,
    }


# -------------------------------------------------------------------
# Writer
# -------------------------------------------------------------------
CATEGORY_COLORS = {
    "breaking_change":       "FFCCCC",  # light red
    "deprecation_notice":    "FFD9B3",  # light orange
    "schema_change":         "FFF2CC",  # light yellow
    "security_notice":       "F4CCCC",  # pink
    "new_endpoint":          "D9EAD3",  # light green
    "policy_change":         "D0E0E3",  # light teal
    "regional_availability": "CFE2F3",  # light blue
    "informational":         "EFEFEF",  # light grey
    "other":                 "FFFFFF",  # white
}

IMPACT_COLORS = {
    "High": "C00000",  # dark red
    "Med":  "BF8F00",  # amber
    "Low":  "808080",  # grey
    "None": "BFBFBF",  # light grey
}


def write_xlsx(items: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()

    # ---- Sheet 1: full detail ----
    ws = wb.active
    ws.title = "All Items"

    headers = ["#", "Title", "Date", "Category", "Secondary Tags", "BP Impact",
               "Rationale", "Link", "Description"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, it in enumerate(items, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=it["title"])
        ws.cell(row=row, column=3, value=it["date"])
        ws.cell(row=row, column=4, value=it["category"])
        ws.cell(row=row, column=5, value=it["secondary_tags"])
        ws.cell(row=row, column=6, value=it["bp_impact"])
        ws.cell(row=row, column=7, value=it["rationale"])
        ws.cell(row=row, column=8, value=it["link"])
        ws.cell(row=row, column=9, value=it["description"])

        # colorize category cell
        color = CATEGORY_COLORS.get(it["category"], "FFFFFF")
        ws.cell(row=row, column=4).fill = PatternFill("solid", start_color=color)

        # impact cell bold + colored text
        impact_cell = ws.cell(row=row, column=6)
        impact_cell.font = Font(bold=True, color=IMPACT_COLORS.get(it["bp_impact"], "000000"))
        impact_cell.alignment = Alignment(horizontal="center")

        for col in range(1, 10):
            ws.cell(row=row, column=col).alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="center" if col in (1, 3, 6) else "left",
            )

    # column widths
    widths = [5, 55, 22, 22, 22, 12, 40, 60, 80]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 32

    # freeze header
    ws.freeze_panes = "A2"

    # autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(items)+1}"

    # ---- Sheet 2: summary counts ----
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Category counts"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A3"] = "Category"
    summary["B3"] = "Count"
    summary["C3"] = "% of total"
    for c in (summary["A3"], summary["B3"], summary["C3"]):
        c.font = Font(bold=True)

    cats = ["breaking_change", "deprecation_notice", "schema_change", "security_notice",
            "new_endpoint", "policy_change", "regional_availability", "informational", "other"]
    from collections import Counter
    cnt = Counter(it["category"] for it in items)
    total = len(items)
    for i, cat in enumerate(cats, start=4):
        summary.cell(row=i, column=1, value=cat)
        summary.cell(row=i, column=1).fill = PatternFill("solid", start_color=CATEGORY_COLORS.get(cat, "FFFFFF"))
        summary.cell(row=i, column=2, value=cnt.get(cat, 0))
        summary.cell(row=i, column=3, value=f"=B{i}/{total}")
        summary.cell(row=i, column=3).number_format = "0.0%"
    total_row = 4 + len(cats)
    summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    summary.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})").font = Font(bold=True)

    # BP-impact roll-up
    summary["E1"] = "BP-impact distribution"
    summary["E1"].font = Font(bold=True, size=14)
    summary["E3"] = "Impact"
    summary["F3"] = "Count"
    summary["G3"] = "% of total"
    for c in (summary["E3"], summary["F3"], summary["G3"]):
        c.font = Font(bold=True)
    icnt = Counter(it["bp_impact"] for it in items)
    for i, imp in enumerate(["High", "Med", "Low", "None"], start=4):
        summary.cell(row=i, column=5, value=imp)
        summary.cell(row=i, column=5).font = Font(bold=True, color=IMPACT_COLORS.get(imp, "000000"))
        summary.cell(row=i, column=6, value=icnt.get(imp, 0))
        summary.cell(row=i, column=7, value=f"=F{i}/{total}")
        summary.cell(row=i, column=7).number_format = "0.0%"

    for col_letter, w in zip("ABCDEFG", [25, 10, 12, 4, 10, 10, 12]):
        summary.column_dimensions[col_letter].width = w

    # ---- Sheet 3: BP-Impact = High only (the ones ops actually needs to watch) ----
    high = wb.create_sheet("BP-Impact = High")
    for col_idx, h in enumerate(headers, 1):
        c = high.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="C00000")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hrow = 2
    for i, it in enumerate(items, start=1):
        if it["bp_impact"] != "High":
            continue
        high.cell(row=hrow, column=1, value=i)
        high.cell(row=hrow, column=2, value=it["title"])
        high.cell(row=hrow, column=3, value=it["date"])
        high.cell(row=hrow, column=4, value=it["category"])
        high.cell(row=hrow, column=4).fill = PatternFill("solid",
                    start_color=CATEGORY_COLORS.get(it["category"], "FFFFFF"))
        high.cell(row=hrow, column=5, value=it["secondary_tags"])
        high.cell(row=hrow, column=6, value=it["bp_impact"])
        high.cell(row=hrow, column=6).font = Font(bold=True, color=IMPACT_COLORS["High"])
        high.cell(row=hrow, column=7, value=it["rationale"])
        high.cell(row=hrow, column=8, value=it["link"])
        high.cell(row=hrow, column=9, value=it["description"])
        for col in range(1, 10):
            high.cell(row=hrow, column=col).alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="center" if col in (1, 3, 6) else "left",
            )
        hrow += 1
    for col_idx, w in enumerate(widths, 1):
        high.column_dimensions[get_column_letter(col_idx)].width = w
    high.row_dimensions[1].height = 32
    high.freeze_panes = "A2"
    if hrow > 2:
        high.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{hrow-1}"

    wb.save(path)


# -------------------------------------------------------------------
# Main
# -------------------------------------------------------------------
def main():
    text = SRC.read_text(encoding="utf-8")
    items = parse_items(text)
    print(f"parsed {len(items)} items")

    for it in items:
        it.update(classify(it))

    write_xlsx(items, OUT_XLSX)
    print(f"wrote {OUT_XLSX}")

    # print summary
    from collections import Counter
    cnt = Counter(it["category"] for it in items)
    icnt = Counter(it["bp_impact"] for it in items)
    print("\n-- Category distribution --")
    for cat, n in cnt.most_common():
        print(f"  {cat:<25} {n:>4}  ({100*n/len(items):.1f}%)")
    print("\n-- BP-impact distribution --")
    for imp in ["High", "Med", "Low", "None"]:
        n = icnt.get(imp, 0)
        print(f"  {imp:<25} {n:>4}  ({100*n/len(items):.1f}%)")


if __name__ == "__main__":
    main()