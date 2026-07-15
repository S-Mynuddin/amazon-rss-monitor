import feedparser
from datetime import datetime
import json
import sys
import re
import requests
import os
from pathlib import Path
from collections import Counter

# Excel handling
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠️ openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================
RSS_URL = "http://developer-docs.amazon.com/sp-api/changelog.rss"
STATE_FILE = "rss_state.json"
OUTPUT_EXCEL = "amazon_rss_update.xlsx"

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/rss+xml, application/xml, text/xml, */*',
    'Accept-Language': 'en-US,en;q=0.9',
}

# ============================================================
# RSS FETCHING FUNCTIONS (from your original code)
# ============================================================
def fetch_rss_feed():
    """Fetch and parse the RSS feed using requests"""
    print("📡 Fetching RSS feed from Amazon...")
    
    try:
        response = requests.get(RSS_URL, headers=HEADERS, timeout=30)
        print(f"    Status: {response.status_code}")
        
        if response.status_code != 200:
            print(f"❌ HTTP Error: {response.status_code}")
            return None
        
        print(f"    Downloaded: {len(response.content):,} bytes")
        feed = feedparser.parse(response.content)
        
        if feed.bozo:
            print(f"    Warning: {feed.bozo_exception}")
        
        if feed.entries:
            print(f"✅ Success! Found {len(feed.entries)} items")
            return feed
        else:
            print("    No entries found")
            return None
            
    except Exception as e:
        print(f"❌ Error: {e}")
        return None

def fetch_rss_https():
    """Fallback: Try HTTPS with SSL verification disabled"""
    print("\n🔄 Trying HTTPS fallback...")
    https_url = "https://developer-docs.amazon.com/sp-api/changelog.rss"
    
    try:
        response = requests.get(https_url, headers=HEADERS, verify=False, timeout=30)
        if response.status_code == 200:
            feed = feedparser.parse(response.content)
            if feed.entries:
                print(f"✅ HTTPS success! Found {len(feed.entries)} items")
                return feed
        return None
    except Exception as e:
        print(f"❌ HTTPS fallback failed: {e}")
        return None

# ============================================================
# STATE MANAGEMENT
# ============================================================
def load_state():
    """Load previously seen GUIDs from state file"""
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                print(f"📂 Loaded state: {len(data.get('guids', []))} previously seen GUIDs")
                return data.get('guids', [])
        except Exception as e:
            print(f"⚠️ Error loading state file: {e}")
            return []
    else:
        print("ℹ️ No state file found. This is the first run.")
        return []

def save_state(guids, total_items, last_build_date):
    """Save current GUIDs to state file"""
    try:
        data = {
            "last_check": datetime.now().isoformat(),
            "last_build_date": last_build_date,
            "total_items": total_items,
            "guids": guids
        }
        with open(STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"💾 State saved: {len(guids)} GUIDs")
        return True
    except Exception as e:
        print(f"❌ Error saving state: {e}")
        return False

def extract_guid(entry):
    """Extract GUID from RSS entry (handles different formats)"""
    guid = entry.get('id') or entry.get('guid') or entry.get('link')
    if hasattr(guid, 'value'):
        guid = guid.value
    if not guid:
        guid = entry.get('link', '')
    return str(guid)

# ============================================================
# CLASSIFICATION ENGINE (from client's script)
# ============================================================
# Compiled regex patterns for classification
BREAKING_RX = re.compile(
    r"\b(will (no longer|be removed|end|be sunset|be turned off|stop (working|being)|be replaced by)|"
    r"no longer offer|no longer support|will cease|is (being )?(retired|removed|sunset)|"
    r"final release|breaking change|action required|must migrate|"
    r"end of (support|life)|will be deprecated|final version|"
    r"will not be available|shall no longer)",
    re.I,
)

DEPRECATION_RX = re.compile(
    r"\b(deprecat(ed|ion|ing)|sunset(ting)?|retiring|will be retired|"
    r"end[- ]of[- ]life|EOL|removal of|removed from the api|will retire)",
    re.I,
)

SCHEMA_RX = re.compile(
    r"\b(attribute usage|enumeration values?|product type definitions?|"
    r"listing attribute|new (field|attribute|parameter|column|response|value|enum|metric|property|dataset|event|list|group-by|element|report type|analytics report)|"
    r"additional (field|attribute|parameter|response|value|metric)|"
    r"schema( change| update)?|response format|payload change|title structure|"
    r"XSD|response body|"
    r"json schema|extended (with|to include)|includes? new|expand(s|ed|ing) (the )?(schema|response|attributes)|"
    r"(now |will )(include|includes|contain|contains|return|returns|support|supports|accept|accepts|show|shows|surface|surfaces|has|have)|"
    r"response now|operation now|(API|report|dataset|feed|endpoint) now|"
    r"format (change|update|standardization)|"
    r"new fee (type|typ)|new value(s)? for|"
    r"rate limit (change|update|increase|adjust)|throttling adjust|"
    r"(attribute|field|property|parameter|object) (added|is now (optional|required)|has been added)|"
    r"(will require|requires|required for|no longer requires)|"
    r"has been added to|"
    r"adds the |adds a new |"
    r"validation of|validate |"
    r"clarification on|"
    r"upcoming (product type|attribute))",
    re.I,
)

SECURITY_RX = re.compile(
    r"\b(OAuth|LWA|restricted data token|\bRDT\b|PII|"
    r"authentication|access token|refresh token|credential|"
    r"grantless|IAM|AWS Signature|encryption|"
    r"permissions? (change|required|update)|authorization|"
    r"security (update|advisory|patch|change)|role scope|application role|"
    r"vulnerabilit(y|ies)|CVE-\d|Log4j|"
    r"escaping issue|injection)",
    re.I,
)

NEW_ENDPOINT_RX = re.compile(
    r"\b(launch(ed|ing)|introduc(ing|es)|announcing|"
    r"new (API|operation|endpoint|version|capability|SDK|sample|documentation|dashboard|tool|workflow)|"
    r"general availability|\bGA release\b|now available|"
    r"prebuilt|new: |Health Dashboard|"
    r"enhancement(s)? to (the )?[A-Z][A-Za-z ]+API|"
    r"Send to Amazon will replace|will replace the|"
    r"released|beta|preview|v\d{4}-\d{2}(-\d{2})?|v\d+_\d+|"
    r"is (now )?live|goes live|rollout|roll[- ]out)",
    re.I,
)

POLICY_RX = re.compile(
    r"\b(policy|policies|terms of service|terms and conditions|"
    r"seller (agreement|policy|policies|performance)|"
    r"(FBA|FBM) (requirements?|fees?)|handling time (requirements?|changes?)|"
    r"prep (services?|requirements?|policy)|labeling (services?|requirements?)|"
    r"fee (change|schedule|update|structure)|surcharge|"
    r"compliance|regulatory|requirement (change|update))",
    re.I,
)

REGION_RX = re.compile(
    r"\b(EU marketplace|EU stores|EU sellers|EU FBM|"
    r"United Kingdom|Germany|France|Italy|Spain|Netherlands|Belgium|Poland|Sweden|Turkey|"
    r"Japan|India|Australia|Brazil|Mexico|Singapore|United Arab Emirates|Saudi Arabia|"
    r"South Africa|Egypt|"
    r"marketplace(s)?|region(al)?|country|countries)",
    re.I,
)

INFORMATIONAL_RX = re.compile(
    r"\b(release notes|welcome|documentation site|"
    r"developer newsletter|monthly (update|release|digest)|"
    r"SP-API Release|Selling Partner API Release|Weekly (Update|Digest)|"
    r"quarterly (update|release)|announcement digest|update digest|"
    r"documentation update|blog post|case study)",
    re.I,
)

BP_HIGH_RX = re.compile(
    r"\b(FBA|Fulfillment by Amazon|inventory|restock|replenish|fulfillment|"
    r"Reserved (FC|Customer|Staging)|FC Transfer|"
    r"sales report|orders? report|order( |s )?(api|report)|"
    r"product type|listings? item|catalog item|"
    r"feeds? api|shipment (api|report)|"
    r"inbound (api|shipment)|receiving)",
    re.I,
)

BP_MED_RX = re.compile(
    r"\b(reports? api|pricing (api|report)|notifications|subscription|"
    r"grantless|application management|"
    r"data kiosk|selling partner (api|insights))",
    re.I,
)

BP_LOW_RX = re.compile(
    r"\b(brand analytics|A\+ content|advertising|promotions|solicitations|"
    r"messaging|customer solicitations|invoices?|"
    r"tax(es|ation)|VAT|GST|"
    r"vendor central|1P|first-?party vendor|"
    r"amazon business|B2B|"
    r"eBook|Kindle|Prime Video|automotive fitment|apparel|"
    r"fashion|kitchen|beauty)",
    re.I,
)

BP_MARKETS = re.compile(r"\b(US|USA|United States|CA|Canada|North America|global|worldwide|all marketplaces)\b", re.I)
NON_BP_MARKETS = re.compile(
    r"\b(EU|Europe|UK|United Kingdom|Germany|France|Italy|Spain|Netherlands|Belgium|Poland|Sweden|Turkey|"
    r"Japan|India|Australia|AU|Brazil|Mexico|Singapore|UAE|United Arab Emirates|Saudi Arabia|"
    r"South Africa|Egypt|Middle East)\b",
    re.I,
)

# Color schemes for Excel
CATEGORY_COLORS = {
    "breaking_change":       "FFCCCC",  # light red
    "deprecation_notice":    "FFD9B3",  # light orange
    "schema_change":         "FFF2CC",  # light yellow
    "security_notice":       "F4CCCC",  # pink
    "new_endpoint":          "D9EAD3",  # light green
    "policy_change":         "D0E0E3",  # light teal
    "regional_availability": "CFE2F3",  # light blue
    "informational":         "EFEFEF",  # light grey
    "other":                 "FFFFFF",  # white
}

IMPACT_COLORS = {
    "High": "C00000",  # dark red
    "Med":  "BF8F00",  # amber
    "Low":  "808080",  # grey
    "None": "BFBFBF",  # light grey
}

def classify_item(title: str, description: str) -> dict:
    """Classify a single item using the client's rules"""
    haystack = f"{title}\n{description}"
    
    # Primary category (priority-ordered)
    if BREAKING_RX.search(haystack):
        primary = "breaking_change"
    elif DEPRECATION_RX.search(haystack):
        primary = "deprecation_notice"
    elif SECURITY_RX.search(haystack):
        primary = "security_notice"
    elif SCHEMA_RX.search(haystack):
        primary = "schema_change"
    elif NEW_ENDPOINT_RX.search(haystack) and not INFORMATIONAL_RX.search(haystack):
        primary = "new_endpoint"
    elif POLICY_RX.search(haystack):
        primary = "policy_change"
    elif INFORMATIONAL_RX.search(haystack):
        primary = "informational"
    elif REGION_RX.search(haystack):
        primary = "regional_availability"
    else:
        primary = "other"
    
    # Secondary tags
    secondary = []
    if primary != "breaking_change" and BREAKING_RX.search(haystack):
        secondary.append("breaking")
    if primary != "deprecation_notice" and DEPRECATION_RX.search(haystack):
        secondary.append("deprecation")
    if primary != "schema_change" and SCHEMA_RX.search(haystack):
        secondary.append("schema")
    if primary != "security_notice" and SECURITY_RX.search(haystack):
        secondary.append("security")
    if primary != "new_endpoint" and NEW_ENDPOINT_RX.search(haystack):
        secondary.append("new")
    if primary != "policy_change" and POLICY_RX.search(haystack):
        secondary.append("policy")
    if primary != "regional_availability" and REGION_RX.search(haystack):
        secondary.append("regional")
    
    # BP Impact rating
    has_bp_market = bool(BP_MARKETS.search(haystack))
    has_non_bp_only = bool(NON_BP_MARKETS.search(haystack)) and not has_bp_market
    is_bp_high = bool(BP_HIGH_RX.search(haystack))
    is_bp_med = bool(BP_MED_RX.search(haystack))
    is_bp_low = bool(BP_LOW_RX.search(haystack))
    
    if has_non_bp_only and not is_bp_high:
        bp_impact = "None"
    elif is_bp_high:
        bp_impact = "High"
    elif is_bp_med:
        bp_impact = "Med"
    elif is_bp_low:
        bp_impact = "Low"
    elif primary in ("breaking_change", "deprecation_notice"):
        bp_impact = "Med"
    elif primary == "informational":
        bp_impact = "Low"
    else:
        bp_impact = "Med"
    
    # Rationale
    parts = [f"→ {primary}"]
    if secondary:
        parts.append(f"(+ {', '.join(secondary)})")
    if has_non_bp_only:
        parts.append("[non-BP market only]")
    if is_bp_high:
        parts.append("[BP-domain: FBA/inventory/listings]")
    elif is_bp_med:
        parts.append("[BP-domain: reports/notifications]")
    elif is_bp_low:
        parts.append("[BP-domain: ads/brand/vendor — not BP]")
    rationale = " ".join(parts)
    
    return {
        "category": primary,
        "secondary_tags": ", ".join(secondary),
        "bp_impact": bp_impact,
        "rationale": rationale
    }

# ============================================================
# FIND NEW ITEMS
# ============================================================
def find_new_items(feed, old_guids):
    """Find items that are in the feed but NOT in old_guids"""
    current_guids = []
    new_items = []
    
    for entry in feed.entries:
        guid = extract_guid(entry)
        current_guids.append(guid)
        
        if guid not in old_guids:
            new_items.append({
                'guid': guid,
                'title': entry.get('title', 'No Title'),
                'link': entry.get('link', ''),
                'pub_date': entry.get('published', 'Unknown date'),
                'description': entry.get('description', 'No description'),
                'author': entry.get('author', ''),
                'raw_entry': entry
            })
    
    return current_guids, new_items

# ============================================================
# EXCEL BUILDER
# ============================================================
def build_excel_file(feed, new_items, current_guids):
    """Build Excel file with Part 1 (New Items) and Part 2 (All Items classified)"""
    
    wb = openpyxl.Workbook()
    
    # ============================================================
    # SHEET 1: PART 1 - NEW UPDATES
    # ============================================================
    ws_new = wb.active
    ws_new.title = "Part 1 - New Updates"
    
    # Headers
    headers = ["#", "Title", "Date", "Category", "Secondary Tags", "BP Impact", 
               "Rationale", "Link", "Description"]
    
    for col_idx, h in enumerate(headers, 1):
        c = ws_new.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add new items
    for idx, item in enumerate(new_items, 1):
        row = idx + 1
        
        # Classify the item
        classification = classify_item(item['title'], item['description'])
        
        ws_new.cell(row=row, column=1, value=idx)
        ws_new.cell(row=row, column=2, value=item['title'])
        ws_new.cell(row=row, column=3, value=item['pub_date'])
        ws_new.cell(row=row, column=4, value=classification['category'])
        ws_new.cell(row=row, column=5, value=classification['secondary_tags'])
        ws_new.cell(row=row, column=6, value=classification['bp_impact'])
        ws_new.cell(row=row, column=7, value=classification['rationale'])
        ws_new.cell(row=row, column=8, value=item['link'])
        ws_new.cell(row=row, column=9, value=item['description'][:500] + "..." if len(item['description']) > 500 else item['description'])
        
        # Color category cell
        color = CATEGORY_COLORS.get(classification['category'], "FFFFFF")
        ws_new.cell(row=row, column=4).fill = PatternFill("solid", start_color=color)
        
        # Color impact cell
        impact_cell = ws_new.cell(row=row, column=6)
        impact_cell.font = Font(bold=True, color=IMPACT_COLORS.get(classification['bp_impact'], "000000"))
        impact_cell.alignment = Alignment(horizontal="center")
    
    # Auto-fit columns
    widths = [5, 55, 22, 22, 22, 12, 40, 60, 80]
    for col_idx, w in enumerate(widths, 1):
        ws_new.column_dimensions[get_column_letter(col_idx)].width = w
    ws_new.row_dimensions[1].height = 32
    ws_new.freeze_panes = "A2"
    
    # ============================================================
    # SHEET 2: PART 2 - ALL ITEMS (Classified)
    # ============================================================
    ws_all = wb.create_sheet("Part 2 - All Items")
    
    # Headers
    for col_idx, h in enumerate(headers, 1):
        c = ws_all.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add all items with classification
    for idx, entry in enumerate(feed.entries, 1):
        row = idx + 1
        guid = extract_guid(entry)
        is_new = guid in [item['guid'] for item in new_items]
        
        title = entry.get('title', 'No Title')
        if is_new:
            title = "⭐ " + title  # Mark new items
        
        classification = classify_item(entry.get('title', ''), entry.get('description', ''))
        
        ws_all.cell(row=row, column=1, value=idx)
        ws_all.cell(row=row, column=2, value=title)
        ws_all.cell(row=row, column=3, value=entry.get('published', 'Unknown date'))
        ws_all.cell(row=row, column=4, value=classification['category'])
        ws_all.cell(row=row, column=5, value=classification['secondary_tags'])
        ws_all.cell(row=row, column=6, value=classification['bp_impact'])
        ws_all.cell(row=row, column=7, value=classification['rationale'])
        ws_all.cell(row=row, column=8, value=entry.get('link', ''))
        desc = entry.get('description', '')
        ws_all.cell(row=row, column=9, value=desc[:500] + "..." if len(desc) > 500 else desc)
        
        # Color category cell
        color = CATEGORY_COLORS.get(classification['category'], "FFFFFF")
        ws_all.cell(row=row, column=4).fill = PatternFill("solid", start_color=color)
        
        # Color impact cell
        impact_cell = ws_all.cell(row=row, column=6)
        impact_cell.font = Font(bold=True, color=IMPACT_COLORS.get(classification['bp_impact'], "000000"))
        impact_cell.alignment = Alignment(horizontal="center")
    
    # Auto-fit columns
    for col_idx, w in enumerate(widths, 1):
        ws_all.column_dimensions[get_column_letter(col_idx)].width = w
    ws_all.row_dimensions[1].height = 32
    ws_all.freeze_panes = "A2"
    
    # ============================================================
    # SHEET 3: SUMMARY STATISTICS
    # ============================================================
    ws_summary = wb.create_sheet("Summary")
    
    # Category counts
    ws_summary["A1"] = "📊 CATEGORY DISTRIBUTION"
    ws_summary["A1"].font = Font(bold=True, size=14)
    
    ws_summary["A3"] = "Category"
    ws_summary["B3"] = "Count"
    ws_summary["C3"] = "% of Total"
    for c in (ws_summary["A3"], ws_summary["B3"], ws_summary["C3"]):
        c.font = Font(bold=True)
    
    # Count categories
    all_items = []
    for entry in feed.entries:
        classification = classify_item(entry.get('title', ''), entry.get('description', ''))
        all_items.append(classification)
    
    cat_count = Counter(item['category'] for item in all_items)
    total = len(all_items)
    
    row = 4
    for cat, count in cat_count.most_common():
        ws_summary.cell(row=row, column=1, value=cat)
        ws_summary.cell(row=row, column=1).fill = PatternFill("solid", start_color=CATEGORY_COLORS.get(cat, "FFFFFF"))
        ws_summary.cell(row=row, column=2, value=count)
        ws_summary.cell(row=row, column=3, value=f"=B{row}/{total}")
        ws_summary.cell(row=row, column=3).number_format = "0.0%"
        row += 1
    
    # BP Impact distribution
    ws_summary["E1"] = "📊 BP-IMPACT DISTRIBUTION"
    ws_summary["E1"].font = Font(bold=True, size=14)
    
    ws_summary["E3"] = "Impact"
    ws_summary["F3"] = "Count"
    ws_summary["G3"] = "% of Total"
    for c in (ws_summary["E3"], ws_summary["F3"], ws_summary["G3"]):
        c.font = Font(bold=True)
    
    impact_count = Counter(item['bp_impact'] for item in all_items)
    
    row = 4
    for impact in ["High", "Med", "Low", "None"]:
        count = impact_count.get(impact, 0)
        ws_summary.cell(row=row, column=5, value=impact)
        ws_summary.cell(row=row, column=5).font = Font(bold=True, color=IMPACT_COLORS.get(impact, "000000"))
        ws_summary.cell(row=row, column=6, value=count)
        ws_summary.cell(row=row, column=7, value=f"=F{row}/{total}")
        ws_summary.cell(row=row, column=7).number_format = "0.0%"
        row += 1
    
    # New items count
    ws_summary["A10"] = f"🆕 New Items Found: {len(new_items)}"
    ws_summary["A10"].font = Font(bold=True, size=12, color="0066CC")
    ws_summary["A11"] = f"📄 Total Items in Feed: {total}"
    ws_summary["A11"].font = Font(bold=True, size=12)
    ws_summary["A12"] = f"📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A12"].font = Font(size=11)
    
    # Set column widths
    for col_letter, w in zip("ABCDEFG", [25, 12, 14, 4, 20, 12, 14]):
        ws_summary.column_dimensions[col_letter].width = w
    
    # ============================================================
    # SHEET 4: BP-IMPACT = HIGH (Critical items)
    # ============================================================
    ws_high = wb.create_sheet("Critical - BP Impact High")
    
    # Headers
    for col_idx, h in enumerate(headers, 1):
        c = ws_high.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="C00000")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row = 2
    for idx, entry in enumerate(feed.entries, 1):
        classification = classify_item(entry.get('title', ''), entry.get('description', ''))
        
        if classification['bp_impact'] != "High":
            continue
        
        guid = extract_guid(entry)
        is_new = guid in [item['guid'] for item in new_items]
        title = entry.get('title', 'No Title')
        if is_new:
            title = "⭐ " + title
        
        ws_high.cell(row=row, column=1, value=idx)
        ws_high.cell(row=row, column=2, value=title)
        ws_high.cell(row=row, column=3, value=entry.get('published', 'Unknown date'))
        ws_high.cell(row=row, column=4, value=classification['category'])
        ws_high.cell(row=row, column=4).fill = PatternFill("solid", start_color=CATEGORY_COLORS.get(classification['category'], "FFFFFF"))
        ws_high.cell(row=row, column=5, value=classification['secondary_tags'])
        ws_high.cell(row=row, column=6, value=classification['bp_impact'])
        ws_high.cell(row=row, column=6).font = Font(bold=True, color=IMPACT_COLORS["High"])
        ws_high.cell(row=row, column=7, value=classification['rationale'])
        ws_high.cell(row=row, column=8, value=entry.get('link', ''))
        desc = entry.get('description', '')
        ws_high.cell(row=row, column=9, value=desc[:500] + "..." if len(desc) > 500 else desc)
        
        for col in range(1, 10):
            ws_high.cell(row=row, column=col).alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="center" if col in (1, 3, 6) else "left",
            )
        row += 1
    
    for col_idx, w in enumerate(widths, 1):
        ws_high.column_dimensions[get_column_letter(col_idx)].width = w
    ws_high.row_dimensions[1].height = 32
    ws_high.freeze_panes = "A2"
    
    # Save the workbook
    wb.save(OUTPUT_EXCEL)
    print(f"✅ Excel file saved: {OUTPUT_EXCEL}")
    return len(new_items), total

# ============================================================
# MAIN FUNCTION
# ============================================================
def main():
    print("\n" + "=" * 80)
    print("📊 AMAZON RSS FEED MONITOR WITH CLASSIFICATION")
    print("=" * 80)
    
    # Fetch RSS feed
    feed = fetch_rss_feed()
    if feed is None or not feed.entries:
        feed = fetch_rss_https()
    
    if feed is None or not feed.entries:
        print("\n❌ Failed to fetch RSS feed.")
        return
    
    # Load previous state
    old_guids = load_state()
    
    # Find new items
    current_guids, new_items = find_new_items(feed, old_guids)
    
    print(f"\n📊 Current feed: {len(current_guids)} items")
    print(f"📊 Previously seen: {len(old_guids)} items")
    print(f"🆕 New items found: {len(new_items)}")
    
    if new_items:
        print("\n🆕 New items:")
        for item in new_items:
            print(f"   • {item['title'][:80]}...")
    
    # Build Excel file
    print("\n📝 Building Excel file...")
    new_count, total_count = build_excel_file(feed, new_items, current_guids)
    
    # Save state
    print("\n💾 Updating state...")
    last_build_date = feed.feed.get('lastBuildDate', datetime.now().isoformat())
    save_state(current_guids, len(current_guids), last_build_date)
    
    # Summary
    print("\n" + "=" * 80)
    print("✅ COMPLETED")
    print("=" * 80)
    print(f"\n📁 Files created:")
    print(f"   - {OUTPUT_EXCEL} (Excel report with classification)")
    print(f"   - {STATE_FILE} (Tracking state)")
    print(f"\n📊 Summary:")
    print(f"   - New items: {new_count}")
    print(f"   - Total items: {total_count}")
    print(f"   - Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("\n📧 Ready to email to client: Attach the Excel file")
    print("\n" + "=" * 80)

if __name__ == "__main__":
    # Suppress SSL warnings
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    main()