"""
Amazon SP-API RSS Monitor - Streamlit Web App
Deployed on Streamlit Cloud
"""

import feedparser
from datetime import datetime
import json
import re
import requests
import os
from pathlib import Path
from collections import Counter
import pandas as pd
import streamlit as st
import io
import base64
import tempfile

# Excel handling
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

# ============================================================
# PAGE CONFIGURATION
# ============================================================
st.set_page_config(
    page_title="Amazon SP-API RSS Monitor",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# CONFIGURATION
# ============================================================
RSS_URL = "http://developer-docs.amazon.com/sp-api/changelog.rss"
STATE_FILE = "rss_state.json"

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/rss+xml, application/xml, text/xml, */*',
    'Accept-Language': 'en-US,en;q=0.9',
}

# Initialize session state
if 'initialized' not in st.session_state:
    st.session_state.initialized = True
    st.session_state.feed_data = None
    st.session_state.new_items = []
    st.session_state.last_update = None

# ============================================================
# RSS FETCHING FUNCTIONS
# ============================================================
@st.cache_data(ttl=3600)
def fetch_rss_feed():
    """Fetch and parse the RSS feed using requests"""
    try:
        response = requests.get(RSS_URL, headers=HEADERS, timeout=30)
        
        if response.status_code != 200:
            return None
        
        feed = feedparser.parse(response.content)
        
        if feed.entries:
            return feed
        else:
            return None
            
    except Exception as e:
        return None

# ============================================================
# STATE MANAGEMENT
# ============================================================
def load_state():
    """Load previously seen GUIDs from state file"""
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('guids', [])
        except Exception as e:
            return []
    else:
        return []

def save_state(guids, total_items, last_build_date):
    """Save current GUIDs to state file"""
    try:
        data = {
            "last_check": datetime.now().isoformat(),
            "last_build_date": last_build_date,
            "total_items": total_items,
            "guids": guids
        }
        with open(STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        return False

def extract_guid(entry):
    """Extract GUID from RSS entry"""
    guid = entry.get('id') or entry.get('guid') or entry.get('link')
    if hasattr(guid, 'value'):
        guid = guid.value
    if not guid:
        guid = entry.get('link', '')
    return str(guid)

# ============================================================
# CLASSIFICATION ENGINE
# ============================================================
BREAKING_RX = re.compile(
    r"\b(will (no longer|be removed|end|be sunset|be turned off|stop (working|being)|be replaced by)|"
    r"no longer offer|no longer support|will cease|is (being )?(retired|removed|sunset)|"
    r"final release|breaking change|action required|must migrate|"
    r"end of (support|life)|will be deprecated|final version|"
    r"will not be available|shall no longer)",
    re.I,
)

DEPRECATION_RX = re.compile(
    r"\b(deprecat(ed|ion|ing)|sunset(ting)?|retiring|will be retired|"
    r"end[- ]of[- ]life|EOL|removal of|removed from the api|will retire)",
    re.I,
)

SCHEMA_RX = re.compile(
    r"\b(attribute usage|enumeration values?|product type definitions?|"
    r"listing attribute|new (field|attribute|parameter|column|response|value|enum|metric|property|dataset|event|list|group-by|element|report type|analytics report)|"
    r"additional (field|attribute|parameter|response|value|metric)|"
    r"schema( change| update)?|response format|payload change|title structure|"
    r"XSD|response body|"
    r"json schema|extended (with|to include)|includes? new|expand(s|ed|ing) (the )?(schema|response|attributes)|"
    r"(now |will )(include|includes|contain|contains|return|returns|support|supports|accept|accepts|show|shows|surface|surfaces|has|have)|"
    r"response now|operation now|(API|report|dataset|feed|endpoint) now|"
    r"format (change|update|standardization)|"
    r"new fee (type|typ)|new value(s)? for|"
    r"rate limit (change|update|increase|adjust)|throttling adjust|"
    r"(attribute|field|property|parameter|object) (added|is now (optional|required)|has been added)|"
    r"(will require|requires|required for|no longer requires)|"
    r"has been added to|"
    r"adds the |adds a new |"
    r"validation of|validate |"
    r"clarification on|"
    r"upcoming (product type|attribute))",
    re.I,
)

SECURITY_RX = re.compile(
    r"\b(OAuth|LWA|restricted data token|\bRDT\b|PII|"
    r"authentication|access token|refresh token|credential|"
    r"grantless|IAM|AWS Signature|encryption|"
    r"permissions? (change|required|update)|authorization|"
    r"security (update|advisory|patch|change)|role scope|application role|"
    r"vulnerabilit(y|ies)|CVE-\d|Log4j|"
    r"escaping issue|injection)",
    re.I,
)

NEW_ENDPOINT_RX = re.compile(
    r"\b(launch(ed|ing)|introduc(ing|es)|announcing|"
    r"new (API|operation|endpoint|version|capability|SDK|sample|documentation|dashboard|tool|workflow)|"
    r"general availability|\bGA release\b|now available|"
    r"prebuilt|new: |Health Dashboard|"
    r"enhancement(s)? to (the )?[A-Z][A-Za-z ]+API|"
    r"Send to Amazon will replace|will replace the|"
    r"released|beta|preview|v\d{4}-\d{2}(-\d{2})?|v\d+_\d+|"
    r"is (now )?live|goes live|rollout|roll[- ]out)",
    re.I,
)

POLICY_RX = re.compile(
    r"\b(policy|policies|terms of service|terms and conditions|"
    r"seller (agreement|policy|policies|performance)|"
    r"(FBA|FBM) (requirements?|fees?)|handling time (requirements?|changes?)|"
    r"prep (services?|requirements?|policy)|labeling (services?|requirements?)|"
    r"fee (change|schedule|update|structure)|surcharge|"
    r"compliance|regulatory|requirement (change|update))",
    re.I,
)

REGION_RX = re.compile(
    r"\b(EU marketplace|EU stores|EU sellers|EU FBM|"
    r"United Kingdom|Germany|France|Italy|Spain|Netherlands|Belgium|Poland|Sweden|Turkey|"
    r"Japan|India|Australia|Brazil|Mexico|Singapore|United Arab Emirates|Saudi Arabia|"
    r"South Africa|Egypt|"
    r"marketplace(s)?|region(al)?|country|countries)",
    re.I,
)

INFORMATIONAL_RX = re.compile(
    r"\b(release notes|welcome|documentation site|"
    r"developer newsletter|monthly (update|release|digest)|"
    r"SP-API Release|Selling Partner API Release|Weekly (Update|Digest)|"
    r"quarterly (update|release)|announcement digest|update digest|"
    r"documentation update|blog post|case study)",
    re.I,
)

BP_HIGH_RX = re.compile(
    r"\b(FBA|Fulfillment by Amazon|inventory|restock|replenish|fulfillment|"
    r"Reserved (FC|Customer|Staging)|FC Transfer|"
    r"sales report|orders? report|order( |s )?(api|report)|"
    r"product type|listings? item|catalog item|"
    r"feeds? api|shipment (api|report)|"
    r"inbound (api|shipment)|receiving)",
    re.I,
)

BP_MED_RX = re.compile(
    r"\b(reports? api|pricing (api|report)|notifications|subscription|"
    r"grantless|application management|"
    r"data kiosk|selling partner (api|insights))",
    re.I,
)

BP_LOW_RX = re.compile(
    r"\b(brand analytics|A\+ content|advertising|promotions|solicitations|"
    r"messaging|customer solicitations|invoices?|"
    r"tax(es|ation)|VAT|GST|"
    r"vendor central|1P|first-?party vendor|"
    r"amazon business|B2B|"
    r"eBook|Kindle|Prime Video|automotive fitment|apparel|"
    r"fashion|kitchen|beauty)",
    re.I,
)

BP_MARKETS = re.compile(r"\b(US|USA|United States|CA|Canada|North America|global|worldwide|all marketplaces)\b", re.I)
NON_BP_MARKETS = re.compile(
    r"\b(EU|Europe|UK|United Kingdom|Germany|France|Italy|Spain|Netherlands|Belgium|Poland|Sweden|Turkey|"
    r"Japan|India|Australia|AU|Brazil|Mexico|Singapore|UAE|United Arab Emirates|Saudi Arabia|"
    r"South Africa|Egypt|Middle East)\b",
    re.I,
)

CATEGORY_COLORS = {
    "breaking_change":       "FFCCCC",
    "deprecation_notice":    "FFD9B3",
    "schema_change":         "FFF2CC",
    "security_notice":       "F4CCCC",
    "new_endpoint":          "D9EAD3",
    "policy_change":         "D0E0E3",
    "regional_availability": "CFE2F3",
    "informational":         "EFEFEF",
    "other":                 "FFFFFF",
}

def classify_item(title: str, description: str) -> dict:
    """Classify a single item"""
    haystack = f"{title}\n{description}"
    
    if BREAKING_RX.search(haystack):
        primary = "breaking_change"
    elif DEPRECATION_RX.search(haystack):
        primary = "deprecation_notice"
    elif SECURITY_RX.search(haystack):
        primary = "security_notice"
    elif SCHEMA_RX.search(haystack):
        primary = "schema_change"
    elif NEW_ENDPOINT_RX.search(haystack) and not INFORMATIONAL_RX.search(haystack):
        primary = "new_endpoint"
    elif POLICY_RX.search(haystack):
        primary = "policy_change"
    elif INFORMATIONAL_RX.search(haystack):
        primary = "informational"
    elif REGION_RX.search(haystack):
        primary = "regional_availability"
    else:
        primary = "other"
    
    secondary = []
    if primary != "breaking_change" and BREAKING_RX.search(haystack):
        secondary.append("breaking")
    if primary != "deprecation_notice" and DEPRECATION_RX.search(haystack):
        secondary.append("deprecation")
    if primary != "schema_change" and SCHEMA_RX.search(haystack):
        secondary.append("schema")
    if primary != "security_notice" and SECURITY_RX.search(haystack):
        secondary.append("security")
    if primary != "new_endpoint" and NEW_ENDPOINT_RX.search(haystack):
        secondary.append("new")
    if primary != "policy_change" and POLICY_RX.search(haystack):
        secondary.append("policy")
    if primary != "regional_availability" and REGION_RX.search(haystack):
        secondary.append("regional")
    
    has_bp_market = bool(BP_MARKETS.search(haystack))
    has_non_bp_only = bool(NON_BP_MARKETS.search(haystack)) and not has_bp_market
    is_bp_high = bool(BP_HIGH_RX.search(haystack))
    is_bp_med = bool(BP_MED_RX.search(haystack))
    is_bp_low = bool(BP_LOW_RX.search(haystack))
    
    if has_non_bp_only and not is_bp_high:
        bp_impact = "None"
    elif is_bp_high:
        bp_impact = "High"
    elif is_bp_med:
        bp_impact = "Med"
    elif is_bp_low:
        bp_impact = "Low"
    elif primary in ("breaking_change", "deprecation_notice"):
        bp_impact = "Med"
    elif primary == "informational":
        bp_impact = "Low"
    else:
        bp_impact = "Med"
    
    parts = [f"→ {primary}"]
    if secondary:
        parts.append(f"(+ {', '.join(secondary)})")
    if has_non_bp_only:
        parts.append("[non-BP market only]")
    if is_bp_high:
        parts.append("[BP-domain: FBA/inventory/listings]")
    elif is_bp_med:
        parts.append("[BP-domain: reports/notifications]")
    elif is_bp_low:
        parts.append("[BP-domain: ads/brand/vendor — not BP]")
    rationale = " ".join(parts)
    
    return {
        "category": primary,
        "secondary_tags": ", ".join(secondary),
        "bp_impact": bp_impact,
        "rationale": rationale
    }

# ============================================================
# EXCEL BUILDER (FIXED)
# ============================================================
def build_excel_file(feed, new_items):
    """Build Excel file with classifications"""
    wb = openpyxl.Workbook()
    
    # SHEET 1: New Items
    ws_new = wb.active
    ws_new.title = "Part 1 - New Updates"
    headers = ["#", "Title", "Date", "Category", "Secondary Tags", "BP Impact", "Rationale", "Link", "Description"]
    
    for col_idx, h in enumerate(headers, 1):
        c = ws_new.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", start_color="FF1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for idx, item in enumerate(new_items, 1):
        row = idx + 1
        classification = classify_item(item['title'], item['description'])
        
        ws_new.cell(row=row, column=1, value=idx)
        ws_new.cell(row=row, column=2, value=item['title'])
        ws_new.cell(row=row, column=3, value=item['pub_date'])
        ws_new.cell(row=row, column=4, value=classification['category'])
        ws_new.cell(row=row, column=5, value=classification['secondary_tags'])
        ws_new.cell(row=row, column=6, value=classification['bp_impact'])
        ws_new.cell(row=row, column=7, value=classification['rationale'])
        ws_new.cell(row=row, column=8, value=item['link'])
        desc = item['description']
        ws_new.cell(row=row, column=9, value=desc[:500] + "..." if len(desc) > 500 else desc)
        
        color = CATEGORY_COLORS.get(classification['category'], "FFFFFF")
        ws_new.cell(row=row, column=4).fill = PatternFill("solid", start_color=f"FF{color}")
    
    widths = [5, 55, 22, 22, 22, 12, 40, 60, 80]
    for col_idx, w in enumerate(widths, 1):
        ws_new.column_dimensions[get_column_letter(col_idx)].width = w
    ws_new.row_dimensions[1].height = 32
    ws_new.freeze_panes = "A2"
    
    # SHEET 2: All Items
    ws_all = wb.create_sheet("Part 2 - All Items")
    for col_idx, h in enumerate(headers, 1):
        c = ws_all.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", start_color="FF2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for idx, entry in enumerate(feed.entries, 1):
        row = idx + 1
        guid = extract_guid(entry)
        is_new = guid in [item['guid'] for item in new_items]
        title = entry.get('title', 'No Title')
        if is_new:
            title = "⭐ " + title
        
        classification = classify_item(entry.get('title', ''), entry.get('description', ''))
        
        ws_all.cell(row=row, column=1, value=idx)
        ws_all.cell(row=row, column=2, value=title)
        ws_all.cell(row=row, column=3, value=entry.get('published', 'Unknown date'))
        ws_all.cell(row=row, column=4, value=classification['category'])
        ws_all.cell(row=row, column=5, value=classification['secondary_tags'])
        ws_all.cell(row=row, column=6, value=classification['bp_impact'])
        ws_all.cell(row=row, column=7, value=classification['rationale'])
        ws_all.cell(row=row, column=8, value=entry.get('link', ''))
        desc = entry.get('description', '')
        ws_all.cell(row=row, column=9, value=desc[:500] + "..." if len(desc) > 500 else desc)
        
        color = CATEGORY_COLORS.get(classification['category'], "FFFFFF")
        ws_all.cell(row=row, column=4).fill = PatternFill("solid", start_color=f"FF{color}")
    
    for col_idx, w in enumerate(widths, 1):
        ws_all.column_dimensions[get_column_letter(col_idx)].width = w
    ws_all.row_dimensions[1].height = 32
    ws_all.freeze_panes = "A2"
    
    # SHEET 3: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "📊 CATEGORY DISTRIBUTION"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Category"
    ws_summary["B3"] = "Count"
    ws_summary["C3"] = "% of Total"
    for c in (ws_summary["A3"], ws_summary["B3"], ws_summary["C3"]):
        c.font = Font(bold=True)
    
    all_items = []
    for entry in feed.entries:
        classification = classify_item(entry.get('title', ''), entry.get('description', ''))
        all_items.append(classification)
    
    cat_count = Counter(item['category'] for item in all_items)
    total = len(all_items)
    row = 4
    for cat, count in cat_count.most_common():
        ws_summary.cell(row=row, column=1, value=cat)
        color = CATEGORY_COLORS.get(cat, "FFFFFF")
        ws_summary.cell(row=row, column=1).fill = PatternFill("solid", start_color=f"FF{color}")
        ws_summary.cell(row=row, column=2, value=count)
        ws_summary.cell(row=row, column=3, value=f"=B{row}/{total}")
        ws_summary.cell(row=row, column=3).number_format = "0.0%"
        row += 1
    
    ws_summary["E1"] = "📊 BP-IMPACT DISTRIBUTION"
    ws_summary["E1"].font = Font(bold=True, size=14)
    ws_summary["E3"] = "Impact"
    ws_summary["F3"] = "Count"
    ws_summary["G3"] = "% of Total"
    for c in (ws_summary["E3"], ws_summary["F3"], ws_summary["G3"]):
        c.font = Font(bold=True)
    
    impact_count = Counter(item['bp_impact'] for item in all_items)
    row = 4
    for impact in ["High", "Med", "Low", "None"]:
        count = impact_count.get(impact, 0)
        ws_summary.cell(row=row, column=5, value=impact)
        ws_summary.cell(row=row, column=5).font = Font(bold=True)
        ws_summary.cell(row=row, column=6, value=count)
        ws_summary.cell(row=row, column=7, value=f"=F{row}/{total}")
        ws_summary.cell(row=row, column=7).number_format = "0.0%"
        row += 1
    
    ws_summary["A10"] = f"🆕 New Items Found: {len(new_items)}"
    ws_summary["A10"].font = Font(bold=True, size=12)
    ws_summary["A11"] = f"📄 Total Items in Feed: {total}"
    ws_summary["A11"].font = Font(bold=True, size=12)
    ws_summary["A12"] = f"📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A12"].font = Font(size=11)
    
    for col_letter, w in zip("ABCDEFG", [25, 12, 14, 4, 20, 12, 14]):
        ws_summary.column_dimensions[col_letter].width = w
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============================================================
# FIND NEW ITEMS
# ============================================================
def find_new_items(feed, old_guids):
    """Find items that are in the feed but NOT in old_guids"""
    current_guids = []
    new_items = []
    
    for entry in feed.entries:
        guid = extract_guid(entry)
        current_guids.append(guid)
        
        if guid not in old_guids:
            new_items.append({
                'guid': guid,
                'title': entry.get('title', 'No Title'),
                'link': entry.get('link', ''),
                'pub_date': entry.get('published', 'Unknown date'),
                'description': entry.get('description', 'No description'),
                'author': entry.get('author', ''),
                'raw_entry': entry
            })
    
    return current_guids, new_items

# ============================================================
# STREAMLIT UI
# ============================================================
def main():
    st.title("📊 Amazon SP-API RSS Monitor")
    st.markdown("*Automated Changelog Monitoring with Classification*")
    
    with st.sidebar:
        st.header("⚙️ Controls")
        
        if st.button("🔄 Refresh Feed", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
        
        st.divider()
        st.header("📊 Statistics")
        
        if st.session_state.feed_data:
            feed = st.session_state.feed_data
            total_items = len(feed.entries)
            new_count = len(st.session_state.new_items)
            
            st.metric("Total Items", total_items)
            st.metric("New Items", new_count, delta=new_count if new_count > 0 else None)
            
            if st.session_state.last_update:
                st.caption(f"Last updated: {st.session_state.last_update}")
        else:
            st.info("Click 'Refresh Feed' to load data")
        
        st.divider()
        st.header("ℹ️ About")
        st.caption("This app monitors the Amazon SP-API changelog RSS feed and classifies each update by:")
        st.caption("• Category (breaking_change, deprecation_notice, etc.)")
        st.caption("• BP Impact (High, Med, Low, None)")
        st.caption("Updates marked with ⭐ are new since your last check.")
    
    if st.button("📥 Fetch Latest Feed", use_container_width=True, type="primary"):
        with st.spinner("Fetching RSS feed from Amazon..."):
            feed = fetch_rss_feed()
            
            if feed and feed.entries:
                st.session_state.feed_data = feed
                old_guids = load_state()
                current_guids, new_items = find_new_items(feed, old_guids)
                st.session_state.new_items = new_items
                st.session_state.last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                last_build_date = feed.feed.get('lastBuildDate', datetime.now().isoformat())
                save_state(current_guids, len(current_guids), last_build_date)
                
                st.success(f"✅ Feed loaded! Found {len(feed.entries)} items, {len(new_items)} new.")
                st.rerun()
            else:
                st.error("❌ Failed to fetch RSS feed. Please try again.")
    
    if st.session_state.feed_data:
        feed = st.session_state.feed_data
        new_items = st.session_state.new_items
        
        tab1, tab2, tab3, tab4 = st.tabs([
            "🆕 New Updates",
            "📋 All Items",
            "📊 Summary",
            "⬇️ Download Excel"
        ])
        
        with tab1:
            st.subheader(f"🆕 New Updates ({len(new_items)} items)")
            if new_items:
                for idx, item in enumerate(new_items, 1):
                    classification = classify_item(item['title'], item['description'])
                    
                    with st.expander(f"{idx}. {item['title'][:80]}...", expanded=(idx <= 3)):
                        col1, col2 = st.columns([3, 1])
                        with col1:
                            st.write(f"**Title:** {item['title']}")
                            st.write(f"**Date:** {item['pub_date']}")
                            st.write(f"**Link:** [{item['link']}]({item['link']})")
                        with col2:
                            color = CATEGORY_COLORS.get(classification['category'], "#FFFFFF")
                            st.markdown(f"**Category:** <span style='background-color:{color};padding:2px 8px;border-radius:4px;'>{classification['category']}</span>", unsafe_allow_html=True)
                            st.write(f"**BP Impact:** {classification['bp_impact']}")
                            st.write(f"**Secondary:** {classification['secondary_tags']}")
                        
                        st.write(f"**Rationale:** {classification['rationale']}")
                        st.write(f"**Description:** {item['description'][:300]}...")
            else:
                st.info("No new items found. Everything is up to date! ✅")
        
        with tab2:
            st.subheader(f"📋 All Items ({len(feed.entries)} items)")
            
            data = []
            for entry in feed.entries:
                guid = extract_guid(entry)
                is_new = guid in [item['guid'] for item in new_items]
                classification = classify_item(entry.get('title', ''), entry.get('description', ''))
                
                data.append({
                    "#": len(data) + 1,
                    "Title": ("⭐ " if is_new else "") + entry.get('title', 'No Title')[:60],
                    "Date": entry.get('published', 'Unknown date'),
                    "Category": classification['category'],
                    "BP Impact": classification['bp_impact'],
                    "Secondary Tags": classification['secondary_tags'],
                    "Link": entry.get('link', '')
                })
            
            df = pd.DataFrame(data)
            st.dataframe(df, use_container_width=True, height=400)
        
        with tab3:
            st.subheader("📊 Summary Statistics")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Category Distribution**")
                cat_data = []
                for entry in feed.entries:
                    classification = classify_item(entry.get('title', ''), entry.get('description', ''))
                    cat_data.append(classification['category'])
                cat_counts = Counter(cat_data)
                
                cat_df = pd.DataFrame({
                    'Category': list(cat_counts.keys()),
                    'Count': list(cat_counts.values())
                }).sort_values('Count', ascending=False)
                st.dataframe(cat_df, use_container_width=True)
            
            with col2:
                st.write("**BP Impact Distribution**")
                impact_data = []
                for entry in feed.entries:
                    classification = classify_item(entry.get('title', ''), entry.get('description', ''))
                    impact_data.append(classification['bp_impact'])
                impact_counts = Counter(impact_data)
                
                impact_df = pd.DataFrame({
                    'Impact': list(impact_counts.keys()),
                    'Count': list(impact_counts.values())
                }).sort_values('Count', ascending=False)
                st.dataframe(impact_df, use_container_width=True)
            
            st.divider()
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Items", len(feed.entries))
            with col2:
                st.metric("New Items", len(new_items))
            with col3:
                high_impact = sum(1 for item in [classify_item(e.get('title', ''), e.get('description', '')) for e in feed.entries] if item['bp_impact'] == 'High')
                st.metric("High Impact", high_impact)
            with col4:
                breaking = sum(1 for item in [classify_item(e.get('title', ''), e.get('description', '')) for e in feed.entries] if item['category'] == 'breaking_change')
                st.metric("Breaking Changes", breaking)
        
        with tab4:
            st.subheader("⬇️ Download Excel Report")
            
            if st.button("📊 Generate Excel File", use_container_width=True):
                with st.spinner("Generating Excel file..."):
                    excel_data = build_excel_file(feed, new_items)
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=excel_data,
                        file_name=f"amazon_rss_update_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                    st.success("✅ Excel file ready for download!")
            
            st.info("The Excel file contains 4 sheets:\n\n- **Part 1 - New Updates**: Only new items with classifications\n- **Part 2 - All Items**: Complete feed with classifications (⭐ marks new items)\n- **Summary**: Statistics and distribution charts\n- **Critical - BP Impact High**: Filtered view of high-impact items")
    else:
        st.info("👈 Click 'Fetch Latest Feed' in the sidebar to get started!")
        
        st.markdown("""
        ### 🚀 Features
        
        - **Automated RSS Fetching**: Pulls the latest Amazon SP-API changelog
        - **Smart Classification**: Categorizes each update into 9 categories
        - **BP Impact Rating**: Rates items by business relevance (High/Med/Low/None)
        - **Excel Export**: Download categorized data as Excel file
        - **State Tracking**: Remembers what you've already seen
        
        ### 📂 Categories
        
        | Category | Description |
        |----------|-------------|
        | breaking_change | Existing behavior will stop working |
        | deprecation_notice | Retirement announced, future date |
        | schema_change | Fields/attributes/enums changing |
        | security_notice | Auth/credentials/permissions |
        | new_endpoint | New capability shipped |
        | policy_change | Fees/policy/requirements |
        | regional_availability | Market-specific rollouts |
        | informational | Release notes/documentation |
        | other | Uncategorized |
        """)

if __name__ == "__main__":
    main()